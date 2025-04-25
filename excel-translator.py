import openai
from openpyxl import load_workbook, Workbook
import pandas as pd
import os
import json
from tqdm import tqdm
import re
import logging
import csv

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("translation_log.txt", encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# API 키는 실행 시 입력받습니다

# 지원되는 언어 코드와 이름 매핑
LANGUAGE_CODES = {
    'ko': 'Korean',
    'en': 'English',
    'ja': 'Japanese',
    'zh': 'Chinese',
    'es': 'Spanish',
    'fr': 'French',
    'de': 'German',
    'ru': 'Russian',
    'pt': 'Portuguese',
    'it': 'Italian',
    'ar': 'Arabic',
    'hi': 'Hindi',
    'vi': 'Vietnamese',
    'th': 'Thai',
    'id': 'Indonesian',
    'tr': 'Turkish',
    'nl': 'Dutch',
    'he': 'Hebrew',
    'pl': 'Polish',
    'sv': 'Swedish',
    'fi': 'Finnish',
    'da': 'Danish',
    'no': 'Norwegian',
    'el': 'Greek',
    'hu': 'Hungarian',
    'cs': 'Czech',
    'uk': 'Ukrainian',
    'fa': 'Persian',
    'ro': 'Romanian',
    'bg': 'Bulgarian'
}

# 언어별 문자 범위 정의
LANGUAGE_RANGES = {
    'ko': {'ranges': [('\uac00', '\ud7af')], 'name': 'Korean', 'detect_func': 'has_korean'},  # 한글
    'ja': {'ranges': [('\u3040', '\u30ff'), ('\u3400', '\u4dbf'), ('\u4e00', '\u9fff')], 'name': 'Japanese'},  # 일본어(히라가나, 가타카나, 한자)
    'zh': {'ranges': [('\u4e00', '\u9fff')], 'name': 'Chinese'},  # 중국어 간체/번체
    'ru': {'ranges': [('\u0400', '\u04FF')], 'name': 'Russian'},  # 키릴 문자(러시아어 등)
    'ar': {'ranges': [('\u0600', '\u06FF')], 'name': 'Arabic'},  # 아랍어
    'he': {'ranges': [('\u0590', '\u05FF')], 'name': 'Hebrew'},  # 히브리어
    'th': {'ranges': [('\u0E00', '\u0E7F')], 'name': 'Thai'},  # 태국어
    'hi': {'ranges': [('\u0900', '\u097F')], 'name': 'Hindi'},  # 힌디어(데바나가리)
    'el': {'ranges': [('\u0370', '\u03FF')], 'name': 'Greek'},  # 그리스어
    # 라틴 문자 기반 언어들은 특별한 감지 방법 사용
}

def has_korean(text):
    if not isinstance(text, str) or not text.strip():
        return False
    # 한글 유니코드 범위
    return bool(re.search('[\uac00-\ud7af]', text))

def has_japanese(text):
    if not isinstance(text, str) or not text.strip():
        return False
    # 일본어 문자(히라가나, 가타카나, 일부 한자) 유니코드 범위
    return bool(re.search('[\u3040-\u30ff\u3400-\u4dbf\u4e00-\u9fff]', text))

def has_specific_language_chars(text, ranges):
    """특정 유니코드 범위의 문자가 텍스트에 포함되어 있는지 확인"""
    if not isinstance(text, str) or not text.strip():
        return False
    
    for start, end in ranges:
        pattern = f'[{start}-{end}]'
        if re.search(pattern, text):
            return True
    return False

def detect_language(text):
    """텍스트의 언어를 감지하여 가장 가능성 높은 언어 코드 반환"""
    if not isinstance(text, str) or not text.strip():
        return None
    
    # 각 언어별 문자 범위 확인
    for lang_code, lang_info in LANGUAGE_RANGES.items():
        if has_specific_language_chars(text, lang_info['ranges']):
            return lang_code
    
    # 특수 언어 확인
    if has_korean(text):
        return 'ko'
    if has_japanese(text):
        return 'ja'
    
    # 라틴 문자 기반 언어는 기본적으로 영어로 간주 (정확한 구분은 어려움)
    if re.search('[a-zA-Z]', text):
        return 'en'
    
    # 기타 문자
    return None

def has_target_language(text, lang_code, target_lang=None):
    """
    지정된 언어가 텍스트에 포함되어 있는지 확인
    모든 주요 언어 감지 지원
    target_lang이 제공되면 해당 언어는 제외
    """
    # 언어 감지 기능이 없으면 단순 텍스트 존재 여부로 판단
    if not isinstance(text, str) or not text.strip():
        return False
    
    # 'any' 코드인 경우, 대상 언어와 다른 언어인지 확인
    if lang_code == 'any' and target_lang:
        # 대상 언어의 텍스트인지 확인
        detected = detect_language(text)
        # 감지된 언어가 대상 언어와 같으면 번역 대상에서 제외
        if detected == target_lang:
            return False
        # 그 외의 모든 경우는 번역 대상에 포함
        return True
    
    # 특수 언어 체크
    if lang_code == 'ko':
        return has_korean(text)
    elif lang_code == 'ja':
        return has_japanese(text)
    # 언어 범위가 정의된 경우
    elif lang_code in LANGUAGE_RANGES:
        return has_specific_language_chars(text, LANGUAGE_RANGES[lang_code]['ranges'])
    # 언어 코드가 'any'인 경우 모든 텍스트 처리 (target_lang 검사 이미 수행됨)
    elif lang_code == 'any':
        return True
    # 기타 언어는 해당 언어로 감지된 경우에만 처리
    else:
        detected = detect_language(text)
        return detected == lang_code or detected is None

def batch_translate(contents, src, tgt, grocery=None, max_retry=3):
    """
    contents: List of strings
    grocery: 현재까지 구축된 용어 사전 (선택적)
    Returns: List of translated strings
    """
    # 입력이 비어있거나 너무 작으면 빈 배열 반환
    if not contents:
        return []
    
    # 이미 번역된 용어 확인 및 제외
    if grocery:
        to_translate = []
        translated_items = []
        indices_map = []  # 원본 인덱스 매핑
        
        for i, text in enumerate(contents):
            if text in grocery:
                translated_items.append((i, grocery[text]))
                logging.info(f"Reusing from grocery: '{text}' -> '{grocery[text]}'")
            else:
                to_translate.append(text)
                indices_map.append(i)
        
        # 모든 항목이 이미 번역되었다면 즉시 반환
        if not to_translate:
            logging.info(f"All {len(contents)} items found in grocery, no API call needed")
            # 원래 순서대로 결과 재구성
            result = [""] * len(contents)
            for idx, translated in translated_items:
                result[idx] = translated
            return result
        
        # 남은 항목만 번역 처리
        logging.info(f"{len(translated_items)}/{len(contents)} items found in grocery, {len(to_translate)} items to translate")
        contents = to_translate
    else:
        indices_map = None
        translated_items = []
    
    # 디버깅용 샘플 로깅
    sample_items = min(5, len(contents))
    logging.info(f"Sample items to translate (first {sample_items}): {contents[:sample_items]}")
    
    example_input = ["직업", "연령", "성별", "현재 이용 영어 학습 서비스"]
    example_output = ["Occupation", "Age", "Gender", "Current English Learning Service Used"]
    
    # 입력이 너무 많으면 나누어 처리 (250개 단위로)
    if len(contents) > 250:
        logging.info(f"Large batch detected ({len(contents)} items), splitting into smaller batches")
        result = []
        
        # 현재까지의 배치 결과를 저장할 local_grocery 초기화
        local_grocery = {} if grocery is None else grocery.copy()
        
        for i in range(0, len(contents), 250):
            batch = contents[i:i+250]
            logging.info(f"Processing sub-batch {i//250+1}/{(len(contents)+249)//250} with {len(batch)} items")
            # 이전 배치의 결과를 포함한 용어 사전 전달
            batch_result = batch_translate(batch, src, tgt, local_grocery, max_retry)
            result.extend(batch_result)
            
            # 새 결과를 local_grocery에 추가
            for j, item in enumerate(batch):
                if batch_result[j].startswith("[ERROR]"):
                    continue  # 오류 항목은 사전에 추가하지 않음
                local_grocery[item] = batch_result[j]
            logging.info(f"Local grocery updated, now contains {len(local_grocery)} items")
        
        # 결과가 indices_map을 통해 원래 위치로 재구성되어야 하는 경우
        if indices_map is not None:
            final_result = [""] * (len(indices_map) + len(translated_items))
            # 미리 번역된 항목 위치에 결과 배치
            for idx, translated in translated_items:
                final_result[idx] = translated
            # API로 번역된 항목 위치에 결과 배치
            for i, orig_idx in enumerate(indices_map):
                final_result[orig_idx] = result[i]
            return final_result
        
        return result
    
    # 실제 번역 처리
    prompt = (
        f"You are given a list of text items from an Excel sheet written in {src}. Your task is to translate each item into {tgt}. "
        f"IMPORTANT INSTRUCTIONS:\n"
        f"- You MUST return only a pure JSON array of translated strings.\n"
        f"- The array MUST contain EXACTLY {len(contents)} items, matching the input count.\n"
        f"- Each item in your output must correspond to the same index in the input array.\n"
        f"- Do NOT add any explanations, numbers, or extra text. Do NOT wrap your answer in triple backticks or any code block. Do NOT return any key names, only the array.\n"
        f"- Make sure repeated source terms are always translated to the same target term for consistency throughout the list.\n"
        f"- If you fail to strictly follow these instructions, your response will be automatically rejected and retried.\n"
        f"- Example input ({len(example_input)} items): {json.dumps(example_input, ensure_ascii=False)}\n"
        f"- Example output (MUST be {len(example_input)} items): {json.dumps(example_output, ensure_ascii=False)}\n\n"
        f"Input list ({len(contents)} items):\n{json.dumps(contents, ensure_ascii=False)}"
    )
    
    for attempt in range(max_retry):
        logging.info(f"Calling OpenAI for batch of {len(contents)} items, attempt {attempt+1}...")
        response = openai.chat.completions.create(
            model="gpt-4.1-2025-04-14",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0
        )
        result = response.choices[0].message.content.strip()
        
        # 디버깅: 응답 일부 로깅 (최대 500자)
        if len(result) > 500:
            logging.info(f"Response preview: {result[:500]}...")
        else:
            logging.info(f"Response: {result}")
            
        # Remove code block wrappers if any
        result = re.sub(r"^```json|^```|```$", "", result.strip(), flags=re.MULTILINE)
        try:
            translated = json.loads(result)
            
            # 길이 체크: 파이썬으로 비교
            if isinstance(translated, list):
                logging.info(f"Parsed JSON array with {len(translated)} items (expected {len(contents)})")
                
                if len(translated) == len(contents):
                    # 결과를 원래 순서로 재구성 (indices_map이 있는 경우)
                    if indices_map is not None:
                        final_result = [""] * (len(indices_map) + len(translated_items))
                        # 미리 번역된 항목 위치에 결과 배치
                        for idx, trans in translated_items:
                            final_result[idx] = trans
                        # API로 번역된 항목 위치에 결과 배치
                        for i, orig_idx in enumerate(indices_map):
                            final_result[orig_idx] = translated[i]
                        logging.info(f"Batch translation succeeded with reordering: {len(final_result)} items.")
                        return final_result
                    
                    logging.info(f"Batch translation succeeded with {len(translated)} items.")
                    return translated
                elif len(translated) > len(contents):
                    logging.warning(f"API returned more items than expected. Truncating to {len(contents)} items.")
                    return translated[:len(contents)]
                else:
                    # 응답이 부족한 경우: 가능한 한 많이 사용하고 나머지는 오류 표시
                    if len(translated) >= len(contents) * 0.9:  # 90% 이상이면 부족한 부분 채우기
                        logging.warning(f"API returned fewer items than expected ({len(translated)}/{len(contents)}). Filling missing items.")
                        missing = len(contents) - len(translated)
                        translated.extend(["[TRANSLATION ERROR]"] * missing)
                        return translated
            
            logging.warning(
                f"Batch translation output mismatch: expected {len(contents)}, got {len(translated) if isinstance(translated, list) else 'not a list'}"
            )
        except Exception as e:
            logging.warning(f"Batch translation parsing failed: {str(e)}")
            
            # 고급 오류 복구 시도: 문자열에서 JSON 배열 패턴 찾기
            try:
                logging.info("Attempting advanced error recovery...")
                # 대괄호로 둘러싸인 배열 찾기
                array_match = re.search(r'\[(.*)\]', result, re.DOTALL)
                if array_match:
                    array_content = array_match.group(0)
                    fixed_json = json.loads(array_content)
                    if isinstance(fixed_json, list):
                        logging.info(f"Recovery successful! Found {len(fixed_json)} items.")
                        if len(fixed_json) == len(contents):
                            return fixed_json
                        elif len(fixed_json) > len(contents):
                            return fixed_json[:len(contents)]
            except:
                logging.warning("Advanced recovery failed")
        
        logging.warning("Retrying batch translation...")
    
    logging.error("Batch translation failed after retries.")
    
    # 모든 재시도 실패 시: 원본 텍스트 반환하고 오류 표시
    logging.warning("Returning original text with error markers as fallback")
    return [f"[ERROR] {text}" for text in contents]

def load_file(input_file):
    """
    파일 확장자에 따라 적절한 로더 사용
    지원 형식: xlsx, csv
    """
    ext = os.path.splitext(input_file)[1].lower()
    
    if ext == '.xlsx':
        logging.info(f"Loading Excel file: {input_file}")
        return load_workbook(input_file), None, 'excel'
    elif ext == '.csv':
        logging.info(f"Loading CSV file: {input_file}")
        # CSV는 pandas로 읽고 데이터프레임으로 처리
        df = pd.read_csv(input_file)
        return None, df, 'csv'
    else:
        raise ValueError(f"Unsupported file format: {ext}. Please use .xlsx or .csv files.")

def save_file(data, output_file, file_type):
    """
    파일 형식에 맞게 저장
    """
    ext = os.path.splitext(output_file)[1].lower()
    
    if file_type == 'excel':
        # Excel 워크북 객체를 저장
        data.save(output_file)
        logging.info(f"Excel file saved: {output_file}")
    elif file_type == 'csv':
        # DataFrame을 CSV로 저장
        data.to_csv(output_file, index=False)
        logging.info(f"CSV file saved: {output_file}")
    
    return output_file

def get_language_name(lang_code):
    """언어 코드로부터 언어 이름 반환"""
    return LANGUAGE_CODES.get(lang_code, lang_code)

def main():
    # API 키 입력 받기
    api_key = input("OpenAI API 키를 입력하세요: ").strip()
    openai.api_key = api_key

    # 파일 경로 입력
    input_file = input("번역할 파일 경로를 입력하세요 (xlsx, csv 지원): ").strip().replace("'", "").replace('"', '')
    if not os.path.isfile(input_file):
        logging.error(f"File not found: {input_file}")
        print("파일을 찾을 수 없습니다.")
        return
    
    # 언어 코드 직접 입력
    print("\n지원되는 언어 코드 예시:")
    print("- ko (한국어), ja (일본어), zh (중국어), en (영어)")
    print("- fr (프랑스어), es (스페인어), de (독일어), ru (러시아어)")
    print("- ar (아랍어), hi (힌디어), th (태국어), vi (베트남어)")
    print("- any (모든 언어 감지, 대상 언어 제외)")
    source_lang = input("원본 언어 코드를 입력하세요 (모든 언어: 'any'): ").strip().lower()
    target_lang = input("대상 언어 코드를 입력하세요: ").strip().lower()
    
    # 시트 이름 번역 여부 확인
    translate_sheet_names = input("시트 이름도 번역하시겠습니까? (y/n): ").strip().lower() == 'y'

    # 선택된 언어 코드에 해당하는 언어 이름 찾기
    source_lang_name = "Any language" if source_lang == 'any' else get_language_name(source_lang)
    target_lang_name = get_language_name(target_lang)
    
    # 출력 파일 이름 설정
    output_file = os.path.splitext(input_file)[0] + f"_translated_{target_lang}{os.path.splitext(input_file)[1]}"

    logging.info(f"Starting translation: {input_file}")
    logging.info(f"Source language: {source_lang} ({source_lang_name}), Target language: {target_lang} ({target_lang_name})")
    logging.info(f"Translate sheet names: {translate_sheet_names}")

    # 파일 로드
    try:
        wb, df, file_type = load_file(input_file)
    except ValueError as e:
        logging.error(str(e))
        print(str(e))
        return
    
    # 용어 사전(grocery) 초기화 - 원본 텍스트를 키로, 번역된 텍스트를 값으로 저장
    translation_grocery = {}
    
    # 용어 사전 로그 파일 설정
    grocery_file = os.path.splitext(input_file)[0] + "_translation_grocery.json"
    
    # 모든 시트/데이터의 총 번역 대상 항목 수와 사전 덕분에 절약한 항목 수 추적
    total_items = 0
    total_saved = 0

    if file_type == 'excel':
        # Excel 파일 처리
        total_sheets = len(wb.worksheets)
        logging.info(f"Found {total_sheets} sheets")
        
        # 시트 이름 번역 (선택 시)
        if translate_sheet_names and total_sheets > 0:
            print("\n시트 이름 번역 중...")
            sheet_names = [ws.title for ws in wb.worksheets]
            sheet_names_to_translate = []
            
            # 번역이 필요한 시트 이름 확인
            for name in sheet_names:
                if has_target_language(name, source_lang, target_lang if source_lang == 'any' else None):
                    if name in translation_grocery:
                        logging.info(f"Sheet name reused from grocery: '{name}' -> '{translation_grocery[name]}'")
                    else:
                        sheet_names_to_translate.append(name)
            
            # 시트 이름 번역
            if sheet_names_to_translate:
                try:
                    translated_names = batch_translate(sheet_names_to_translate, source_lang_name, target_lang_name)
                    
                    # 번역 결과를 grocery에 추가
                    for i, name in enumerate(sheet_names_to_translate):
                        translation_grocery[name] = translated_names[i]
                    
                    # 번역 결과를 적용 - 새 워크북으로 복사하면서 시트 이름 변경
                    new_wb = Workbook()
                    # 기본 시트 제거
                    if new_wb.worksheets:
                        new_wb.remove(new_wb.worksheets[0])
                        
                    # 각 시트 복사하며 이름 변경
                    for i, ws in enumerate(wb.worksheets):
                        old_name = ws.title
                        if has_target_language(old_name, source_lang, target_lang if source_lang == 'any' else None):
                            if old_name in translation_grocery:
                                new_name = translation_grocery[old_name]
                            else:
                                new_name = old_name  # 번역 실패 시 원본 이름 유지
                        else:
                            new_name = old_name  # 대상 언어가 아닌 경우 원본 이름 유지
                            
                        # 유효한 시트 이름으로 변환 (특수문자 제한 등)
                        new_name = new_name[:31]  # Excel 시트 이름 길이 제한
                        new_name = re.sub(r'[\[\]\:\*\?\/\\]', '_', new_name)  # 금지된 문자 대체
                        
                        # 새 시트 생성 및 복사
                        new_ws = new_wb.create_sheet(title=new_name)
                        
                        # 시트 내용 복사
                        for row in ws.iter_rows():
                            for cell in row:
                                new_ws[cell.coordinate] = cell.value
                                
                        print(f"  시트 이름 번역: '{old_name}' -> '{new_name}'")
                    
                    # 기존 워크북 대신 새 워크북 사용
                    wb = new_wb
                    
                except Exception as e:
                    logging.error(f"Error during sheet name translation: {str(e)}")
                    print(f"시트 이름 번역 중 오류 발생: {e}")
        
        # 시트 내용 번역
        for ws_idx, ws in enumerate(wb.worksheets):
            logging.info(f"Processing sheet: {ws.title} ({ws_idx+1}/{total_sheets})")
            print(f"\n[{ws.title}] ({ws_idx+1}/{total_sheets}) -- {source_lang_name} 텍스트 추출 중...")
            targets = []
            grocery_hits = []  # 사전에서 찾은 항목 저장
            
            for row in ws.iter_rows():
                for cell in row:
                    if has_target_language(cell.value, source_lang, target_lang if source_lang == 'any' else None):
                        content = cell.value.strip() if isinstance(cell.value, str) else str(cell.value)
                        # 사전에 있는지 확인
                        if content in translation_grocery:
                            grocery_hits.append({
                                "Sheet": ws.title,
                                "Cell": cell.coordinate,
                                "Content": content,
                                "Translated": translation_grocery[content]
                            })
                        else:
                            targets.append({
                                "Sheet": ws.title,
                                "Cell": cell.coordinate,
                                "Content": content
                            })

            if not targets and not grocery_hits:
                logging.info(f"No {source_lang_name} text found in sheet '{ws.title}'. Skipping.")
                print(f"  No {source_lang_name} text found in [{ws.title}]. Skipping.")
                continue
            
            # 사전 히트 통계
            sheet_total = len(targets) + len(grocery_hits)
            sheet_saved = len(grocery_hits)
            total_items += sheet_total
            total_saved += sheet_saved
            
            logging.info(f"Sheet '{ws.title}': {sheet_total} total items, {sheet_saved} reused from grocery ({sheet_saved/sheet_total*100:.1f}% saved)")
            print(f"  {sheet_total} total items, {sheet_saved} reused from grocery ({sheet_saved/sheet_total*100:.1f}% saved)")
            
            # 사전에서 찾은 항목들은 즉시 적용
            for hit in grocery_hits:
                ws[hit["Cell"]] = hit["Translated"]
            
            # 나머지는 API로 번역
            if targets:
                contents = [item["Content"] for item in targets]
                logging.info(f"Found {len(contents)} items to translate in sheet '{ws.title}'")
                print(f"  {len(contents)} items to translate in sheet '{ws.title}'.")

                try:
                    translated = batch_translate(contents, source_lang_name, target_lang_name, translation_grocery)
                    # 오류 표시 항목 개수 확인
                    error_count = sum(1 for item in translated if isinstance(item, str) and item.startswith("[ERROR]"))
                    if error_count > 0:
                        logging.warning(f"{error_count} items could not be translated and were marked with [ERROR]")
                except Exception as e:
                    logging.error(f"Error during translation on sheet '{ws.title}': {str(e)}")
                    print(f"Error during translation on sheet [{ws.title}]: {e}")
                    return

                # 번역 결과 길이 체크 및 조정 - batch_translate 함수에서 처리되었으므로 여기서는 검증만
                if len(translated) != len(contents):
                    logging.error(f"Critical error: translated length {len(translated)} != contents length {len(contents)}")
                    translated = translated[:len(contents)] if len(translated) > len(contents) else translated + ["[TRANSLATION ERROR]"] * (len(contents) - len(translated))

                # 번역 결과를 사전에 추가
                for i, content in enumerate(contents):
                    translation_grocery[content] = translated[i]
                
                # 사전 상태 저장 (매 시트 처리마다 업데이트)
                with open(grocery_file, "w", encoding="utf-8") as f:
                    json.dump(translation_grocery, f, ensure_ascii=False, indent=2)
                    logging.info(f"Updated translation grocery with {len(translation_grocery)} items")

                # Save translation map as JSON
                json_path = os.path.splitext(input_file)[0] + f"_{ws.title}_translation_map.json"
                
                # targets와 grocery_hits 모두 포함한 전체 맵 생성
                translation_map = [
                    {
                        "Sheet": targets[i]["Sheet"],
                        "Cell": targets[i]["Cell"],
                        "Content": targets[i]["Content"],
                        "Translated": translated[i],
                        "Source": "API"
                    } for i in range(len(contents))
                ] + [
                    {
                        "Sheet": hit["Sheet"],
                        "Cell": hit["Cell"],
                        "Content": hit["Content"],
                        "Translated": hit["Translated"],
                        "Source": "Grocery"
                    } for hit in grocery_hits
                ]
                
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(translation_map, f, ensure_ascii=False, indent=2)
                    logging.info(f"Translation map saved: {json_path}")
                    print(f"  Translation map saved to: {json_path}")

                # Apply translations to Excel
                for idx, t in enumerate(targets):
                    if idx < len(translated):
                        ws[t["Cell"]] = translated[idx]
                    else:
                        ws[t["Cell"]] = "[TRANSLATION ERROR]"
                        logging.error(f"Index error: {idx} is out of range in translated array")
            
            # 사전에서만 찾아서 번역한 경우 맵 저장
            elif grocery_hits:
                json_path = os.path.splitext(input_file)[0] + f"_{ws.title}_translation_map.json"
                translation_map = [
                    {
                        "Sheet": hit["Sheet"],
                        "Cell": hit["Cell"],
                        "Content": hit["Content"],
                        "Translated": hit["Translated"],
                        "Source": "Grocery"
                    } for hit in grocery_hits
                ]
                
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(translation_map, f, ensure_ascii=False, indent=2)
                    logging.info(f"Translation map saved: {json_path}")
                    print(f"  Translation map saved to: {json_path} (100% from grocery)")
        
        # 번역된 Excel 파일 저장
        save_file(wb, output_file, 'excel')
    
    elif file_type == 'csv':
        # CSV 파일 처리
        print(f"\nCSV 파일 처리 중... {source_lang_name} 텍스트 추출 중...")
        targets = []
        grocery_hits = []  # 사전에서 찾은 항목 저장
        
        # 데이터프레임 순회하며 번역 대상 찾기
        for col in df.columns:
            for idx, value in enumerate(df[col]):
                if has_target_language(value, source_lang, target_lang if source_lang == 'any' else None):
                    content = str(value).strip()
                    # 사전에 있는지 확인
                    if content in translation_grocery:
                        grocery_hits.append({
                            "Row": idx,
                            "Column": col,
                            "Content": content,
                            "Translated": translation_grocery[content]
                        })
                    else:
                        targets.append({
                            "Row": idx,
                            "Column": col,
                            "Content": content
                        })
        
        if not targets and not grocery_hits:
            logging.info(f"No {source_lang_name} text found in CSV file. Skipping.")
            print(f"  No {source_lang_name} text found in CSV file. Skipping.")
            return
        
        # 사전 히트 통계
        total_items = len(targets) + len(grocery_hits)
        total_saved = len(grocery_hits)
        
        logging.info(f"CSV file: {total_items} total items, {total_saved} reused from grocery ({total_saved/total_items*100:.1f}% saved)")
        print(f"  {total_items} total items, {total_saved} reused from grocery ({total_saved/total_items*100:.1f}% saved)")
        
        # 사전에서 찾은 항목들은 즉시 적용
        for hit in grocery_hits:
            df.at[hit["Row"], hit["Column"]] = hit["Translated"]
        
        # 나머지는 API로 번역
        if targets:
            contents = [item["Content"] for item in targets]
            logging.info(f"Found {len(contents)} items to translate in CSV file")
            print(f"  {len(contents)} items to translate in CSV file.")

            try:
                translated = batch_translate(contents, source_lang_name, target_lang_name, translation_grocery)
                # 오류 표시 항목 개수 확인
                error_count = sum(1 for item in translated if isinstance(item, str) and item.startswith("[ERROR]"))
                if error_count > 0:
                    logging.warning(f"{error_count} items could not be translated and were marked with [ERROR]")
            except Exception as e:
                logging.error(f"Error during translation on CSV file: {str(e)}")
                print(f"Error during translation on CSV file: {e}")
                return

            # 번역 결과 길이 체크 및 조정 - batch_translate 함수에서 처리되었으므로 여기서는 검증만
            if len(translated) != len(contents):
                logging.error(f"Critical error: translated length {len(translated)} != contents length {len(contents)}")
                translated = translated[:len(contents)] if len(translated) > len(contents) else translated + ["[TRANSLATION ERROR]"] * (len(contents) - len(translated))

            # 번역 결과를 사전에 추가
            for i, content in enumerate(contents):
                translation_grocery[content] = translated[i]
            
            # 사전 상태 저장
            with open(grocery_file, "w", encoding="utf-8") as f:
                json.dump(translation_grocery, f, ensure_ascii=False, indent=2)
                logging.info(f"Updated translation grocery with {len(translation_grocery)} items")

            # Save translation map as JSON
            json_path = os.path.splitext(input_file)[0] + "_translation_map.json"
            
            # targets와 grocery_hits 모두 포함한 전체 맵 생성
            translation_map = [
                {
                    "Row": targets[i]["Row"],
                    "Column": targets[i]["Column"],
                    "Content": targets[i]["Content"],
                    "Translated": translated[i],
                    "Source": "API"
                } for i in range(len(contents))
            ] + [
                {
                    "Row": hit["Row"],
                    "Column": hit["Column"],
                    "Content": hit["Content"],
                    "Translated": hit["Translated"],
                    "Source": "Grocery"
                } for hit in grocery_hits
            ]
            
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(translation_map, f, ensure_ascii=False, indent=2)
                logging.info(f"Translation map saved: {json_path}")
                print(f"  Translation map saved to: {json_path}")

            # Apply translations to DataFrame
            for idx, t in enumerate(targets):
                if idx < len(translated):
                    df.at[t["Row"], t["Column"]] = translated[idx]
                else:
                    df.at[t["Row"], t["Column"]] = "[TRANSLATION ERROR]"
                    logging.error(f"Index error: {idx} is out of range in translated array")
        
        # 번역된 CSV 파일 저장
        save_file(df, output_file, 'csv')
    
    # 최종 통계
    efficiency = (total_saved / total_items * 100) if total_items > 0 else 0
    logging.info(f"Translation complete! Final grocery size: {len(translation_grocery)} items")
    logging.info(f"Efficiency stats: {total_saved}/{total_items} items reused ({efficiency:.1f}% saved)")
    print(f"\nBatch translation complete! Output file: {output_file}")
    print(f"Final grocery size: {len(translation_grocery)} items")
    print(f"Efficiency stats: {total_saved}/{total_items} items reused ({efficiency:.1f}% saved)")

if __name__ == "__main__":
    main()
